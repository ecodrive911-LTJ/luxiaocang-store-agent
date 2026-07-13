# -*- coding: utf-8 -*-
"""
v4: 生成最终调价方案
从v2比价表中提取CONFIRM记录 + 原v2中已通过的(非"需复核")记录
合并后生成最终调价清单
"""
import openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

def get_final_data(path, label):
    """从v2比价表和复核最终判定中合并数据"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["调价方案1_比价表"]

    # 读取复核最终判定
    review_path = path.replace("_v2.xlsx", "_v2_复核最终判定.xlsx")
    wb_r = openpyxl.load_workbook(review_path, data_only=True)
    ws_r = wb_r["复核最终判定"]

    # 建立复核判定字典 (row -> verdict)
    review_verdicts = {}
    for r in range(2, ws_r.max_row + 1):
        verdict = ws_r.cell(r, 1).value
        reason = ws_r.cell(r, 2).value
        lx_name = ws_r.cell(r, 4).value
        lx_price = ws_r.cell(r, 6).value
        cp_name = ws_r.cell(r, 7).value
        cp_price = ws_r.cell(r, 9).value
        # 用品名+价格做key
        key = (str(lx_name), str(lx_price), str(cp_name), str(cp_price))
        review_verdicts[key] = (verdict, reason)

    # 读取v2比价表中所有记录
    all_records = []
    for r in range(2, ws.max_row + 1):
        vals = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        lx_name = vals[2]
        lx_spec = vals[4]
        lx_price = vals[5]
        lx_purchase = vals[6] if ws.max_column >= 6 else None
        cp_name = vals[8]
        cp_spec = vals[9]
        cp_price = vals[10]
        ratio = vals[11]
        adj = vals[12]
        diff = vals[13]
        direction = vals[14]
        loss = vals[15]
        review = vals[16] if ws.max_column >= 16 else None

        key = (str(lx_name), str(lx_price), str(cp_name), str(cp_price))

        if review and "复核" in str(review):
            # 需复核的记录 → 用复核结果
            verdict, reason = review_verdicts.get(key, ("MANUAL", "未找到复核记录"))
            if verdict == "CONFIRM":
                status = "确认可比"
                include = True
            elif verdict == "REJECT":
                status = f"排除: {reason}"
                include = False
            else:
                status = f"待人工: {reason}"
                include = False
        else:
            # 非需复核 → 原样保留
            status = "直接通过"
            include = True

        all_records.append({
            "lx_name": lx_name, "lx_spec": lx_spec, "lx_price": lx_price,
            "lx_purchase": lx_purchase,
            "cp_name": cp_name, "cp_spec": cp_spec, "cp_price": cp_price,
            "ratio": ratio, "adj": adj, "diff": diff,
            "direction": direction, "loss": loss,
            "status": status, "include": include,
        })

    return all_records

def write_final(path, label, records):
    """生成最终调价方案Excel"""
    wb = openpyxl.Workbook()

    # === Sheet 1: 最终调价清单 ===
    ws = wb.active
    ws.title = "最终调价清单"

    headers = ["序号", "本店品名", "本店规格", "本店售价", "本店采购价",
               "竞品品名", "竞品规格", "竞品售价", "比值(竞品/本店)",
               "建议调价", "差额", "方向", "亏本预警", "状态"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    included = [r for r in records if r["include"]]
    excluded = [r for r in records if not r["include"]]

    for ri, r in enumerate(included, 2):
        ws.cell(ri, 1, ri - 1)
        ws.cell(ri, 2, r["lx_name"])
        ws.cell(ri, 3, r["lx_spec"])
        ws.cell(ri, 4, r["lx_price"])
        ws.cell(ri, 5, r["lx_purchase"])
        ws.cell(ri, 6, r["cp_name"])
        ws.cell(ri, 7, r["cp_spec"])
        ws.cell(ri, 8, r["cp_price"])
        ws.cell(ri, 9, r["ratio"])
        ws.cell(ri, 10, r["adj"])
        ws.cell(ri, 11, r["diff"])
        ws.cell(ri, 12, r["direction"])
        ws.cell(ri, 13, r["loss"])
        ws.cell(ri, 14, r["status"])

        for ci in range(1, 15):
            cell = ws.cell(ri, ci)
            cell.border = border
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(
                horizontal="left" if ci in (2, 6) else "center",
                vertical="center", wrap_text=True)

        # 亏本预警标红
        if r["loss"] and "亏" in str(r["loss"]):
            ws.cell(ri, 13).font = Font(name="微软雅黑", size=10, bold=True, color="FF0000")
            ws.cell(ri, 13).fill = PatternFill("solid", fgColor="FFFFCC")
        # 方向颜色
        d = str(r["direction"] or "")
        if "↓" in d:
            ws.cell(ri, 12).font = Font(name="微软雅黑", size=10, bold=True, color="008000")
        elif "↑" in d:
            ws.cell(ri, 12).font = Font(name="微软雅黑", size=10, bold=True, color="FF0000")

    widths = [6, 38, 14, 9, 9, 38, 14, 9, 12, 10, 10, 8, 10, 14]
    for i, w in enumerate(widths, 1):
        ws_out_col = ws.column_dimensions[get_column_letter(i)]
        ws_out_col.width = w
    ws.freeze_panes = "A2"

    # === Sheet 2: 排除记录 ===
    ws2 = wb.create_sheet("排除记录")
    headers2 = ["本店品名", "本店售价", "竞品品名", "竞品售价", "排除原因"]
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(1, ci, h)
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="808080")
        cell.alignment = Alignment(horizontal="center")

    for ri, r in enumerate(excluded, 2):
        ws2.cell(ri, 1, r["lx_name"])
        ws2.cell(ri, 2, r["lx_price"])
        ws2.cell(ri, 3, r["cp_name"])
        ws2.cell(ri, 4, r["cp_price"])
        ws2.cell(ri, 5, r["status"])
        for ci in range(1, 6):
            ws2.cell(ri, ci).border = border
            ws2.cell(ri, ci).font = Font(name="微软雅黑", size=10)
            ws2.cell(ri, ci).alignment = Alignment(
                horizontal="left" if ci in (1, 3, 5) else "center",
                vertical="center", wrap_text=True)
    for i, w in enumerate([38, 9, 38, 9, 30], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 3: 统计摘要 ===
    ws3 = wb.create_sheet("统计摘要")

    # 方向统计
    dir_counts = Counter(str(r["direction"]) for r in included if r["include"])
    loss_count = sum(1 for r in included if r["loss"] and "亏" in str(r["loss"]))

    ws3.cell(1, 1, f"{label} - 最终调价方案统计").font = Font(name="微软雅黑", bold=True, size=14)
    ws3.cell(3, 1, "纳入调价的商品数").font = Font(name="微软雅黑", bold=True)
    ws3.cell(3, 2, len(included))
    ws3.cell(4, 1, "排除商品数").font = Font(name="微软雅黑", bold=True)
    ws3.cell(4, 2, len(excluded))
    ws3.cell(5, 1, "其中:降价").font = Font(name="微软雅黑", color="008000")
    ws3.cell(5, 2, dir_counts.get("↓降价", 0))
    ws3.cell(6, 1, "其中:涨价").font = Font(name="微软雅黑", color="FF0000")
    ws3.cell(6, 2, dir_counts.get("↑涨价", 0))
    ws3.cell(7, 1, "其中:不变").font = Font(name="微软雅黑")
    ws3.cell(7, 2, dir_counts.get("—不变", 0))
    ws3.cell(8, 1, "亏本预警数").font = Font(name="微软雅黑", bold=True, color="FF0000")
    ws3.cell(8, 2, loss_count)

    ws3.cell(10, 1, "调价规则").font = Font(name="微软雅黑", bold=True)
    ws3.cell(11, 1, "鹿小仓售价 = 竞品售价 - 0.05元")
    ws3.cell(12, 1, "若调价后低于采购价 → 标注亏本预警")
    ws3.cell(13, 1, "排除规则: 品牌不同/规格不一致/套装vs单品/品类不同")

    for r in range(1, 14):
        for c in range(1, 3):
            cell = ws3.cell(r, c)
            if cell.font is None or (hasattr(cell.font, 'name') and cell.font.name != "微软雅黑"):
                cell.font = Font(name="微软雅黑", size=10)

    out_name = path.replace("_v2.xlsx", "_最终调价方案.xlsx")
    wb.save(out_name)

    print(f"\n{label}:")
    print(f"  纳入调价: {len(included)} (降价{dir_counts.get('↓降价',0)} 涨价{dir_counts.get('↑涨价',0)} 不变{dir_counts.get('—不变',0)})")
    print(f"  排除: {len(excluded)}")
    print(f"  亏本预警: {loss_count}")
    print(f"  输出: {out_name}")

    return len(included), len(excluded), loss_count

# 处理两份文件
r1 = get_final_data(
    r"C:\Users\13522\Desktop\鹿小仓广安店\调价方案1_鹿小仓广安店vs小柴购_比价表_v2.xlsx",
    "广安店vs小柴购")
n1, e1, l1 = write_final(
    r"C:\Users\13522\Desktop\鹿小仓广安店\调价方案1_鹿小仓广安店vs小柴购_比价表_v2.xlsx",
    "广安店vs小柴购", r1)

r2 = get_final_data(
    r"C:\Users\13522\Desktop\鹿小仓财富店\调价方案1_鹿小仓财富店vs丽晨超市_比价表_v2.xlsx",
    "财富店vs丽晨超市")
n2, e2, l2 = write_final(
    r"C:\Users\13522\Desktop\鹿小仓财富店\调价方案1_鹿小仓财富店vs丽晨超市_比价表_v2.xlsx",
    "财富店vs丽晨超市", r2)

print(f"\n{'='*60}")
print(f"  全部完成")
print(f"{'='*60}")
print(f"广安店: 纳入调价{n1}条, 排除{e1}条, 亏本预警{l1}条")
print(f"财富店: 纳入调价{n2}条, 排除{e2}条, 亏本预警{l2}条")
print(f"合计: 纳入调价{n1+n2}条, 排除{e1+e2}条, 亏本预警{l1+l2}条")
