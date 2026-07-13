# -*- coding: utf-8 -*-
"""
1. 记录用户删除和标注不执行的条目
2. 生成去掉"是否执行"列的干净终稿
3. 同时生成"用户调整记录"sheet
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ============================================================
# 用户调整记录
# ============================================================

adjustments = {
    "广安店vs小柴购": {
        # 被删除的3条（完全是两个商品/品名不对）
        "deleted": [
            {
                "lx_name": "网易严选 全价三拼鸭肉梨犬粮1800g/袋",
                "lx_price": 105,
                "cp_name": "网易严选 山茶植萃香氛洗衣液500g/瓶 强效去污持久留",
                "cp_price": 5.9,
                "direction": "↓降价",
                "diff": -99.15,
                "reason": "完全是两个商品（犬粮vs洗衣液）",
            },
            {
                "lx_name": "曼秀雷敦 护唇天然保润唇膏原味 4g/支",
                "lx_price": 23.5,
                "cp_name": "曼秀雷敦 男士控油抗痘洁面乳150ml",
                "cp_price": 31,
                "direction": "↑涨价",
                "diff": 7.45,
                "reason": "完全是两个商品（润唇膏vs洁面乳）",
            },
            {
                "lx_name": "朗科USB2.0黑旋风闪存盘 黑色小巧加密U盘 64GB 1个",
                "lx_price": 43,
                "cp_name": "正品联想 高速TF内存卡 16/32/64/128GB",
                "cp_price": 25.9,
                "direction": "↓降价",
                "diff": -17.15,
                "reason": "完全是两个商品（U盘vs内存卡）",
            },
        ],
        # 标注"不执行"的12条
        "not_exec": [
            {"lx_name": "杰士邦 玻尿酸避孕套 3只/盒", "lx_price": 59, "cp_name": "杰士邦 3D大颗粒避孕套3只", "cp_price": 16.9, "reason": "同品牌不同系列，价格差异过大"},
            {"lx_name": "清扬 男士去屑洗发水活力运动薄荷 560g/盒", "lx_price": 68, "cp_name": "清扬男士去屑洗发", "cp_price": 500, "reason": "竞品价格异常（500元不合理）"},
            {"lx_name": "杜蕾斯 水润超薄装玻尿酸避孕套安全套 3只/盒 新老包装随机", "lx_price": 33.5, "cp_name": "杜蕾斯 超薄隐feel装避孕套3只", "cp_price": 17.9, "reason": "同品牌不同系列"},
            {"lx_name": "潘婷 PRO-V乳液修护洗发露 750ml/瓶", "lx_price": 83, "cp_name": "施华蔻 多效修护洗发露 400ml", "cp_price": 41.8, "reason": "品牌不同（潘婷vs施华蔻），规格不同"},
            {"lx_name": "心相印 二层抽纸 200抽/袋 新老包装随机", "lx_price": 7.9, "cp_name": "心相印 茶语丝享抽纸3层130抽/包", "cp_price": 3.5, "reason": "同品牌不同规格（2层200抽vs3层130抽）"},
            {"lx_name": "杜蕾斯 水性聚氨酯避孕套001 3只/盒", "lx_price": 108, "cp_name": "杜蕾斯 Love大胆爱吧避孕套3只", "cp_price": 9.9, "reason": "同品牌不同系列（001 vs Love大胆爱吧）"},
            {"lx_name": "冈本 超润滑超薄避孕套 3只/盒", "lx_price": 31.2, "cp_name": "杜蕾斯 超薄隐feel装避孕套3只", "cp_price": 17.9, "reason": "品牌不同（冈本vs杜蕾斯）"},
            {"lx_name": "【1件特惠 到手12瓶】可口可乐 零度无糖可乐碳酸饮料 500ml*12瓶 无整", "lx_price": 54.7, "cp_name": "整箱可口可乐 高听零度无糖可乐汽水330ml*24听", "cp_price": 64, "reason": "套装数量不同（12瓶500ml vs 24听330ml）"},
            {"lx_name": "吉列 剃须刀手动刮胡刀旋转1刀架1刀头男士剃胡 1/支", "lx_price": 16.9, "cp_name": "吉列 锋隐5男士手动剃须刀 1刀头1刀架", "cp_price": 54.5, "reason": "同品牌不同型号（旋转vs锋隐5）"},
            {"lx_name": "【单支装】小鹿妈妈  牙线棒细线牙线弓形剔牙线 独立包装", "lx_price": 0.5, "cp_name": "小鹿妈妈 旺虎经典牙线牙线棒 50支剔牙弓形一次性便", "cp_price": 4.49, "reason": "数量不同（1支vs50支）"},
            {"lx_name": "【尺寸可选】爆款可爱卡皮巴拉毛绒玩偶玩具超大水豚丑萌公仔玩具1个", "lx_price": 45, "cp_name": "尺寸可选 网红卡皮巴拉玩偶带背包可拆卸好龟蜜毛绒玩", "cp_price": 18.5, "reason": "同品但款式/尺寸可能不同，价格差异大"},
            {"lx_name": "润本 积雪草冰沙露 30g/支", "lx_price": 9.5, "cp_name": "润本 紫草护肤油50ml", "cp_price": 18, "reason": "同品牌不同商品（冰沙露vs护肤油）"},
        ],
    },
    "财富店vs丽晨超市": {
        "deleted": [
            {
                "lx_name": "瀚思 叶黄素蒸汽眼罩 规格可选 1片/袋 热敷遮光护眼眼贴学生办公午休神器",
                "lx_price": 1.5,
                "cp_name": "女士中筒袜惠选 中筒袜 均码/双 规格可选 春夏",
                "cp_price": 2.9,
                "direction": "↑涨价",
                "diff": 1.35,
                "reason": "完全是两个商品（蒸汽眼罩vs中筒袜）",
            },
        ],
        "not_exec": [
            {"lx_name": "喜之郎美好时光 芝麻海苔卷 12g/袋", "lx_price": 9, "cp_name": "海苔喜之郎 美好时光海苔原味 4.5g/袋", "cp_price": 5.5, "reason": "重量不同（12g vs 4.5g）"},
            {"lx_name": "美好时光 原味海苔 7.5g/袋", "lx_price": 7.5, "cp_name": "海苔喜之郎 美好时光海苔原味 4.5g/袋", "cp_price": 5.5, "reason": "重量不同（7.5g vs 4.5g）"},
            {"lx_name": "康师傅 红烧牛肉面 方便面 110g/桶", "lx_price": 5.5, "cp_name": "泡面袋装5连包康师傅 红烧牛肉面", "cp_price": 12.8, "reason": "数量不同（1桶vs5连包）"},
            {"lx_name": "威猛先生 柑橘清香洁厕液 900g/瓶", "lx_price": 12, "cp_name": "威猛先生 厨房重油污净xiao去污去渍除垢留香（柑橘清香）", "cp_price": 8.9, "reason": "观察（洁厕液vs厨房重油污净，不同用途）"},
        ],
    },
}

# ============================================================
# 生成干净终稿
# ============================================================

final_files = {
    "广安店vs小柴购": r"C:\Users\13522\Desktop\调价方案1_鹿小仓广安店vs小柴购_比价表_最终调价方案.xlsx",
    "财富店vs丽晨超市": r"C:\Users\13522\Desktop\调价方案1_鹿小仓财富店vs丽晨超市_比价表_最终调价方案.xlsx",
}

for label in ["广安店vs小柴购", "财富店vs丽晨超市"]:
    print(f"\n{'='*60}")
    print(f"  {label} — 生成干净终稿")
    print(f"{'='*60}")

    path = final_files[label]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["最终调价清单"]

    # 读取当前数据（只保留"执行"的行，去掉"是否执行"和"状态"列）
    # 表头: 序号, 本店品名, 本店规格, 本店售价, 本店采购价, 竞品品名, 竞品规格, 竞品售价, 比值, 建议调价, 差额, 方向, 亏本预警, 状态, 是否执行
    # 保留前13列，去掉"状态"(14)和"是否执行"(15)

    exec_rows = []
    not_exec_rows = []
    for r in range(2, ws.max_row + 1):
        row_data = [ws.cell(r, c).value for c in range(1, 14)]  # 前13列
        toggle = ws.cell(r, 15).value if ws.max_column >= 15 else "执行"

        # 跳过空行
        if not row_data[1]:
            continue

        if toggle and "不" in str(toggle):
            not_exec_rows.append(row_data)
        else:
            exec_rows.append(row_data)

    print(f"  执行: {len(exec_rows)}条")
    print(f"  不执行: {len(not_exec_rows)}条")

    # 创建新工作簿
    wb_out = openpyxl.Workbook()

    # === Sheet 1: 最终调价清单 ===
    ws1 = wb_out.active
    ws1.title = "最终调价清单"

    headers = ["序号", "本店品名", "本店规格", "本店售价", "本店采购价",
               "竞品品名", "竞品规格", "竞品售价", "比值(竞品/本店)",
               "建议调价", "差额", "方向", "亏本预警"]
    for ci, h in enumerate(headers, 1):
        cell = ws1.cell(1, ci, h)
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, row_data in enumerate(exec_rows, 2):
        # 重新编序号
        row_data[0] = ri - 1
        for ci, val in enumerate(row_data, 1):
            cell = ws1.cell(ri, ci, val)
            cell.border = border
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(
                horizontal="left" if ci in (2, 6) else "center",
                vertical="center", wrap_text=True)

        # 亏本预警标红
        if row_data[12] and "亏" in str(row_data[12]):
            ws1.cell(ri, 13).font = Font(name="微软雅黑", size=10, bold=True, color="FF0000")
            ws1.cell(ri, 13).fill = PatternFill("solid", fgColor="FFFFCC")
        # 方向颜色
        d = str(row_data[11] or "")
        if "↓" in d:
            ws1.cell(ri, 12).font = Font(name="微软雅黑", size=10, bold=True, color="008000")
        elif "↑" in d:
            ws1.cell(ri, 12).font = Font(name="微软雅黑", size=10, bold=True, color="FF0000")

    widths = [6, 38, 14, 9, 9, 38, 14, 9, 12, 10, 10, 8, 10]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # === Sheet 2: 用户调整记录 ===
    ws2 = wb_out.create_sheet("用户调整记录")

    adj = adjustments[label]

    # Part A: 被删除的条目
    ws2.cell(1, 1, f"{label} — 用户调整记录").font = Font(name="微软雅黑", bold=True, size=14)

    ws2.cell(3, 1, "A. 被删除的条目（完全是两个商品/品名不对）").font = Font(name="微软雅黑", bold=True, size=12, color="C00000")
    del_headers = ["序号", "本店品名", "本店售价", "竞品品名", "竞品售价", "方向", "差额", "删除原因"]
    for ci, h in enumerate(del_headers, 1):
        cell = ws2.cell(4, ci, h)
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = Alignment(horizontal="center")

    for ri, item in enumerate(adj["deleted"], 5):
        ws2.cell(ri, 1, ri - 4)
        ws2.cell(ri, 2, item["lx_name"])
        ws2.cell(ri, 3, item["lx_price"])
        ws2.cell(ri, 4, item["cp_name"])
        ws2.cell(ri, 5, item["cp_price"])
        ws2.cell(ri, 6, item["direction"])
        ws2.cell(ri, 7, item["diff"])
        ws2.cell(ri, 8, item["reason"])
        for ci in range(1, 9):
            ws2.cell(ri, ci).border = border
            ws2.cell(ri, ci).font = Font(name="微软雅黑", size=10)
            ws2.cell(ri, ci).alignment = Alignment(
                horizontal="left" if ci in (2, 4, 8) else "center",
                vertical="center", wrap_text=True)

    # Part B: 标注"不执行"的条目
    start_row = 5 + len(adj["deleted"]) + 2
    ws2.cell(start_row, 1, 'B. 标注不执行的条目（品名相同但规格/系列/数量不同）').font = Font(
        name="微软雅黑", bold=True, size=12, color="FF6600")
    ne_headers = ["序号", "本店品名", "本店售价", "竞品品名", "竞品售价", "不执行原因"]
    for ci, h in enumerate(ne_headers, 1):
        cell = ws2.cell(start_row + 1, ci, h)
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="FF6600")
        cell.alignment = Alignment(horizontal="center")

    for ri, item in enumerate(adj["not_exec"], start_row + 2):
        ws2.cell(ri, 1, ri - start_row - 1)
        ws2.cell(ri, 2, item["lx_name"])
        ws2.cell(ri, 3, item["lx_price"])
        ws2.cell(ri, 4, item["cp_name"])
        ws2.cell(ri, 5, item["cp_price"])
        ws2.cell(ri, 6, item["reason"])
        for ci in range(1, 7):
            ws2.cell(ri, ci).border = border
            ws2.cell(ri, ci).font = Font(name="微软雅黑", size=10)
            ws2.cell(ri, ci).alignment = Alignment(
                horizontal="left" if ci in (2, 4, 6) else "center",
                vertical="center", wrap_text=True)

    for i, w in enumerate([6, 42, 9, 42, 9, 30], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 3: 统计摘要 ===
    ws3 = wb_out.create_sheet("统计摘要")
    dir_counts = Counter(str(r[11]) for r in exec_rows)
    loss_count = sum(1 for r in exec_rows if r[12] and "亏" in str(r[12]))

    ws3.cell(1, 1, f"{label} - 最终调价方案统计").font = Font(name="微软雅黑", bold=True, size=14)
    ws3.cell(3, 1, "执行调价商品数").font = Font(name="微软雅黑", bold=True)
    ws3.cell(3, 2, len(exec_rows))
    ws3.cell(4, 1, "降价").font = Font(name="微软雅黑", color="008000")
    ws3.cell(4, 2, dir_counts.get("↓降价", 0))
    ws3.cell(5, 1, "涨价").font = Font(name="微软雅黑", color="FF0000")
    ws3.cell(5, 2, dir_counts.get("↑涨价", 0))
    ws3.cell(6, 1, "亏本预警").font = Font(name="微软雅黑", bold=True, color="FF0000")
    ws3.cell(6, 2, loss_count)
    ws3.cell(7, 1, "用户删除条目").font = Font(name="微软雅黑", color="C00000")
    ws3.cell(7, 2, len(adj["deleted"]))
    ws3.cell(8, 1, "用户标注不执行").font = Font(name="微软雅黑", color="FF6600")
    ws3.cell(8, 2, len(adj["not_exec"]))

    ws3.cell(10, 1, "调价规则").font = Font(name="微软雅黑", bold=True)
    ws3.cell(11, 1, "鹿小仓售价 = 竞品售价 - 0.05元")
    ws3.cell(12, 1, "若调价后低于采购价 → 标注亏本预警")
    ws3.cell(13, 1, "用户已审核并删除/排除不合适条目")

    for r in range(1, 14):
        for c in range(1, 3):
            cell = ws3.cell(r, c)
            if not cell.font.name or cell.font.name != "微软雅黑":
                cell.font = Font(name="微软雅黑", size=10)

    # 保存（覆盖桌面文件）
    wb_out.save(path)

    print(f"  输出: {path}")
    print(f"  降价: {dir_counts.get('↓降价', 0)} | 涨价: {dir_counts.get('↑涨价', 0)} | 亏本预警: {loss_count}")
    print(f"  用户删除: {len(adj['deleted'])}条 | 用户不执行: {len(adj['not_exec'])}条")

# ============================================================
# 汇总
# ============================================================
print(f"\n{'='*60}")
print(f"  全部完成")
print(f"{'='*60}")
for label in ["广安店vs小柴购", "财富店vs丽晨超市"]:
    adj = adjustments[label]
    print(f"  {label}:")
    print(f"    删除{len(adj['deleted'])}条（完全是两个商品/品名不对）")
    print(f"    不执行{len(adj['not_exec'])}条（规格/系列/数量不同）")
