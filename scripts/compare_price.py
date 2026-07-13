# -*- coding: utf-8 -*-
"""
v2: 规格校验 + 采购价熔断 + 方向箭头
鹿小仓(本店) vs 竞品 比价调价方案1
规则:
  1. 品名模糊匹配(同前)
  2. 规格校验:
     - 提取 克重(g)/容积(ml)/数量(个/把/包/瓶/盒/袋等)
     - 克重差≤10g 或 容积差≤10ml → 可调
     - 超10g/ml → 排除
     - 数量不一致(如1瓶 vs 4瓶) → 排除
     - 套装(×N/箱装/组) vs 单品 → 排除
     - 单位不同(g vs ml) → 排除
  3. 价格偏差熔断: 比值<0.5 或 >2.0 → 标注"需复核"
  4. 调价后低于采购价 → 标注"⚠亏本"
  5. 方向: ↓降价 / ↑涨价 / —不变
"""
import openpyxl, re, sys, math
from collections import defaultdict
import difflib
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 通用函数
# ============================================================

STOPWORDS = {"的","和","与","或","及","为","比","型","装","个",
             "包","袋","瓶","支","盒","箱","片","卷","米",
             "ml","g","kg","cm","mm","plus","版","新款","老款",
             "随机","同款","家用","便携","迷你","小型","大号",
             "中号","小号","高档","豪华","特惠","超值","精选"}

def clean_tokens(s):
    if not s: return set()
    s = str(s).strip()
    s = re.sub(r'[,，。./／\-\(\)（）\[\]【】《》]+', ' ', s)
    s = re.sub(r'\d+\s*(ml|g|kg|cm|mm|个|包|袋|瓶|支|盒|箱|片|卷|米)?', '', s)
    return set(t.strip().lower() for t in s.split()
               if len(t.strip()) >= 2 and t.strip() not in STOPWORDS)

def jaccard(a, b):
    if not a or not b: return 0.0
    inter = len(a & b); union = len(a | b)
    return inter / union if union else 0.0

def seq_ratio(a, b):
    if not a or not b: return 0.0
    return difflib.SequenceMatcher(None, a[:60], b[:60]).ratio()

def safe_float(v):
    try:
        return float(v) if v is not None and str(v).strip() != '' else None
    except:
        return None

# ── 规格解析 ───────────────────────────────────────────────
def parse_spec(spec_str):
    """
    从规格字符串中提取: (weight_g, volume_ml, count, is_set, unit)
    weight_g: 克重 (g/kg)
    volume_ml: 容积 (ml/L)
    count: 数量 (个/把/包/瓶/盒/袋/片/卷/对/双/条/根)
    is_set: 是否套装 (含 ×N / N瓶装 / N包组 / 箱 / 组 等)
    unit: 主单位 ('g','ml','count','unknown')
    """
    if not spec_str:
        return None, None, None, False, None

    s = str(spec_str).strip()
    weight_g = None
    volume_ml = None
    count = None
    is_set = False

    # 套装检测: ×N / Nx / N瓶装 / N包组 / N罐装 / 箱装 / 组合装 / 整箱 / 礼盒
    set_patterns = [
        r'[×x*]\s*\d+',           # ×4, x4, *4
        r'\d+\s*(?:瓶|包|罐|袋|盒|支|个|条|块)\s*装',  # 4瓶装, 6罐装
        r'\d+\s*(?:瓶|包|罐|袋|盒|支|个)\s*组',  # 3瓶组
        r'整箱', r'组合装', r'礼盒', r'套装',
        r'\d+\s*[×x]\s*\d+',       # 500ml×4
    ]
    for pat in set_patterns:
        if re.search(pat, s, re.IGNORECASE):
            is_set = True
            break

    # 提取乘法套装中的数量 (×4, x4, *4)
    m = re.search(r'[×x*]\s*(\d+)', s, re.IGNORECASE)
    if m:
        is_set = True

    # 克重 (g / kg)
    m = re.search(r'(\d+(?:\.\d+)?)\s*(kg|g|克)', s, re.IGNORECASE)
    if m:
        val = float(m.group(1))
        unit = m.group(2).lower()
        weight_g = val * 1000 if unit == 'kg' else val

    # 容积 (ml / L)
    m = re.search(r'(\d+(?:\.\d+)?)\s*(ml|l|升|毫升)', s, re.IGNORECASE)
    if m:
        val = float(m.group(1))
        unit = m.group(2).lower()
        volume_ml = val * 1000 if unit in ('l', '升') else val

    # 数量 (个/把/包/瓶/盒/袋/片/卷/对/双/条/根/罐/桶/杯)
    m = re.search(r'(\d+)\s*(个|把|包|瓶|盒|袋|片|卷|对|双|条|根|罐|桶|杯|枚|块|贴)', s)
    if m:
        count = int(m.group(1))

    # 确定主单位
    if weight_g is not None:
        unit = 'g'
    elif volume_ml is not None:
        unit = 'ml'
    elif count is not None:
        unit = 'count'
    else:
        unit = 'unknown'

    return weight_g, volume_ml, count, is_set, unit

def spec_compatible(spec_a, spec_b):
    """
    判断两个规格是否可比:
    返回 (compatible: bool, reason: str)
    策略: 宽松通过, 严格排除
    - 只有双方规格都明确解析且冲突时才排除
    - 规格无法解析 → 通过(标记规格不明)
    - 套装 → 排除(仅明确×N模式)
    """
    w_a, v_a, c_a, set_a, u_a = parse_spec(spec_a)
    w_b, v_b, c_b, set_b, u_b = parse_spec(spec_b)

    # 套装检测: 任一方明确是套装(×N) → 不可比
    if set_a or set_b:
        return False, "套装/多件装-不可比"

    # 单位不同 → 不可比 (g vs ml, g vs count, etc)
    # 仅当双方单位都明确时才比较
    if u_a != 'unknown' and u_b != 'unknown' and u_a != u_b:
        return False, f"单位不同({u_a} vs {u_b})"

    # 克重校验 (同单位g, 双方都有值)
    if w_a is not None and w_b is not None:
        diff = abs(w_a - w_b)
        if diff > 10:
            return False, f"克重差{diff:.0f}g超限"
    elif v_a is not None and v_b is not None:
        diff = abs(v_a - v_b)
        if diff > 10:
            return False, f"容积差{diff:.0f}ml超限"
    elif c_a is not None and c_b is not None:
        if c_a != c_b:
            return False, f"数量不同({c_a} vs {c_b})"
    # 一方有规格另一方没有 → 通过(不再排除)

    return True, "OK"

# ============================================================
# 匹配+输出 主函数
# ============================================================

def run_compare(lx_path, comp_path, out_path, lx_label, comp_label):
    print(f"\n{'='*60}")
    print(f"  {lx_label} vs {comp_label}")
    print(f"{'='*60}")

    wb_lx = openpyxl.load_workbook(lx_path, data_only=True)
    wb_cp = openpyxl.load_workbook(comp_path, data_only=True)
    ws_lx, ws_cp = wb_lx.active, wb_cp.active

    # ── 检测竞品列结构 ─────────────────────────────────────
    cp_headers = [cell.value for cell in ws_cp[1]]
    # 找售价列: 优先"售价"，其次"售卖单价"
    cp_price_col = None
    for i, h in enumerate(cp_headers):
        if h and '售价' in str(h):
            cp_price_col = i
            break
    if cp_price_col is None:
        for i, h in enumerate(cp_headers):
            if h and '单价' in str(h):
                cp_price_col = i
                break
    print(f"竞品售价列索引: {cp_price_col} ({cp_headers[cp_price_col]})")

    # ── 加载本店(鹿小仓) ───────────────────────────────────
    lx_list = []
    for row in ws_lx.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]: continue
        name = str(row[1]).strip()
        price = safe_float(row[4])
        cost  = safe_float(row[5])
        lx_list.append({
            "seq": row[0], "name": name,
            "cat": (row[2] or "未分类").strip(),
            "spec": row[3], "price": price, "cost": cost,
            "bar": row[8],
            "tokens": clean_tokens(name),
        })
    print(f"本店: {len(lx_list)} 个商品")

    # ── 加载竞品 ───────────────────────────────────────────
    NOISE = {"估配送费","红包立减","扫码领红包","红包满减","红包可用",
             "加购","立即购买","收藏","加入购物车","下单立减",
             "叠加优惠","满减券","领券","返利","优惠券"}
    cp_list = []
    for row in ws_cp.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]: continue
        name = str(row[1]).strip()
        if len(name) < 4: continue
        if any(ns in name for ns in NOISE): continue
        price = safe_float(row[cp_price_col])
        if price is not None and (price <= 0 or price > 2000): continue
        cp_list.append({
            "seq": row[0], "name": name,
            "cat": (row[2] or "未分类").strip(),
            "spec": row[3], "price": price,
            "tokens": clean_tokens(name),
        })
    print(f"竞品(过滤后): {len(cp_list)} 个商品")

    # ── 品类分组 ───────────────────────────────────────────
    lx_by_cat = defaultdict(list)
    cp_by_cat = defaultdict(list)
    for d in lx_list: lx_by_cat[d["cat"]].append(d)
    for d in cp_list: cp_by_cat[d["cat"]].append(d)

    all_cats = sorted(set(lx_by_cat) | set(cp_by_cat))

    # ── 匹配 + 规格校验 ────────────────────────────────────
    JACCARD_THR = 0.12
    SEQ_THR     = 0.45

    results = []
    rejected = []  # 被规格校验排除的

    for cat, lx_items in lx_by_cat.items():
        cp_items = cp_by_cat.get(cat, [])
        if not cp_items: continue

        for lx in lx_items:
            if lx["price"] is None: continue
            best_score, best = 0.0, None
            for cp in cp_items:
                if cp["price"] is None: continue
                j = jaccard(lx["tokens"], cp["tokens"])
                s = seq_ratio(lx["name"], cp["name"])
                if j >= JACCARD_THR or s >= SEQ_THR:
                    score = max(j, s)
                    if score > best_score:
                        best_score = score
                        best = cp

            if best:
                # 品名套装检测 (×N / N瓶装 / N包组 等在品名中)
                name_set_a = bool(re.search(r'[×x*]\s*\d+|\d+\s*(?:瓶|包|罐|袋|盒|支|个)\s*(?:装|组)|整箱|组合装|礼盒|套装', str(lx["name"]), re.IGNORECASE))
                name_set_b = bool(re.search(r'[×x*]\s*\d+|\d+\s*(?:瓶|包|罐|袋|盒|支|个)\s*(?:装|组)|整箱|组合装|礼盒|套装', str(best["name"]), re.IGNORECASE))
                if name_set_a != name_set_b:
                    # 一方品名含套装信息另一方不含 → 可能不是同一SKU
                    rejected.append({
                        "lx_name": lx["name"], "lx_spec": lx["spec"],
                        "cp_name": best["name"], "cp_spec": best["spec"],
                        "lx_price": lx["price"], "cp_price": best["price"],
                        "reason": "品名套装不一致", "cat": cat,
                    })
                    continue

                # 规格校验
                ok, reason = spec_compatible(lx["spec"], best["spec"])
                if not ok:
                    rejected.append({
                        "lx_name": lx["name"], "lx_spec": lx["spec"],
                        "cp_name": best["name"], "cp_spec": best["spec"],
                        "lx_price": lx["price"], "cp_price": best["price"],
                        "reason": reason, "cat": cat,
                    })
                    continue

                cp_p = best["price"]
                lx_p = lx["price"]
                lx_cost = lx["cost"]
                ratio = cp_p / lx_p if lx_p != 0 else None
                adj = round(cp_p - 0.05, 2)
                if adj < 0: adj = 0.0
                diff = round(adj - lx_p, 2)

                # 方向
                if diff < 0:
                    direction = "↓降价"
                elif diff > 0:
                    direction = "↑涨价"
                else:
                    direction = "—不变"

                # 亏本检测
                loss_warning = ""
                if lx_cost is not None and adj < lx_cost:
                    loss_warning = "!!亏本"

                # 价格偏差熔断
                review = ""
                if ratio is not None and (ratio < 0.5 or ratio > 2.0):
                    review = "需复核"

                results.append({
                    "lx_name":  lx["name"], "lx_cat": cat,
                    "lx_spec":  lx["spec"], "lx_price": lx_p,
                    "lx_cost":  lx_cost, "lx_bar": lx["bar"],
                    "cp_name":  best["name"], "cp_spec": best["spec"],
                    "cp_price": cp_p,
                    "ratio": ratio, "adj": adj, "diff": diff,
                    "score": best_score,
                    "direction": direction,
                    "loss": loss_warning,
                    "review": review,
                })

    results.sort(key=lambda x: x["score"], reverse=True)
    print(f"匹配成功(通过规格校验): {len(results)}")
    print(f"被规格校验排除: {len(rejected)}")

    comp_r = [r for r in results if r["ratio"] is not None]
    exp  = sum(1 for r in comp_r if r["ratio"] < 1)
    chp  = sum(1 for r in comp_r if r["ratio"] > 1)
    eq   = sum(1 for r in comp_r if abs(r["ratio"] - 1) < 0.001)
    loss_cnt = sum(1 for r in results if r["loss"])
    rev_cnt  = sum(1 for r in results if r["review"])
    dn = sum(1 for r in results if r["direction"].startswith("↓"))
    up = sum(1 for r in results if r["direction"].startswith("↑"))
    sq = sum(1 for r in results if r["direction"].startswith("—"))

    print(f"  应降价(↓): {dn} | 应涨价(↑): {up} | 不变(—): {sq}")
    print(f"  [!]亏本预警: {loss_cnt} | 需复核: {rev_cnt}")

    # ============================================================
    # 写Excel
    # ============================================================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "调价方案1_比价表"

    thin   = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    HEADERS = [
        "序号", "商品名称(本店)", "品类", "规格(本店)",
        "售价(本店)", "采购价", "条码",
        f"竞品品名({comp_label})", "竞品规格", "竞品售价",
        "比值(竞品÷本店)", "方案1建议价\n(竞品-0.05)",
        "调价差额", "方向", "亏本预警", "复核标记", "匹配分"
    ]
    ws.row_dimensions[1].height = 40
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for ri, r in enumerate(results, 2):
        ratio_s = f"{r['ratio']:.4f}" if r["ratio"] is not None else "—"
        diff_s  = f"{r['diff']:+.2f}" if r["diff"] is not None else "—"
        vals = [
            ri-1, r["lx_name"], r["lx_cat"], r["lx_spec"],
            r["lx_price"], r["lx_cost"], r["lx_bar"],
            r["cp_name"], r["cp_spec"], r["cp_price"],
            ratio_s, r["adj"], diff_s,
            r["direction"], r["loss"], r["review"],
            f"{r['score']:.2f}",
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.border = border
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci in (5,6,10,12) and isinstance(v, (int, float)):
                cell.number_format = '0.00'
            if ci in (2, 8):
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 比值颜色
        if r["ratio"] is not None:
            cell = ws.cell(ri, 11)
            if r["ratio"] < 1:
                cell.font = Font(name="微软雅黑", size=10, color="C00000", bold=True)
            elif r["ratio"] > 1:
                cell.font = Font(name="微软雅黑", size=10, color="375623", bold=True)

        # 差额颜色
        if r["diff"] is not None:
            cell = ws.cell(ri, 13)
            if r["diff"] < 0:
                cell.font = Font(name="微软雅黑", size=10, color="C00000", bold=True)
            elif r["diff"] > 0:
                cell.font = Font(name="微软雅黑", size=10, color="375623", bold=True)

        # 方向颜色
        cell = ws.cell(ri, 14)
        if r["direction"].startswith("↓"):
            cell.font = Font(name="微软雅黑", size=10, color="C00000", bold=True)
        elif r["direction"].startswith("↑"):
            cell.font = Font(name="微软雅黑", size=10, color="375623", bold=True)

        # 亏本预警
        if r["loss"]:
            cell = ws.cell(ri, 15)
            cell.value = "!!亏本"
            cell.font = Font(name="微软雅黑", size=10, color="FF0000", bold=True)
            cell.fill = PatternFill("solid", fgColor="FFF2CC")

        # 复核标记
        if r["review"]:
            cell = ws.cell(ri, 16)
            cell.font = Font(name="微软雅黑", size=10, color="FF6600", bold=True)
            cell.fill = PatternFill("solid", fgColor="FCE4D6")

    WIDTHS = [5, 38, 9, 14, 9, 9, 16, 38, 14, 9, 14, 16, 9, 8, 8, 8, 7]
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── 统计摘要 sheet ──────────────────────────────────────
    ws2 = wb.create_sheet("统计摘要")
    stat = [
        ("指标", "数值"),
        (f"本店({lx_label})商品总数", len(lx_list)),
        (f"竞品({comp_label})商品总数(过滤后)", len(cp_list)),
        ("匹配成功(通过规格校验)", len(results)),
        ("被规格校验排除", len(rejected)),
        ("匹配率", f"{len(results)/len(lx_list)*100:.1f}%" if lx_list else "0%"),
        ("可比价商品数", len(comp_r)),
        ("  本店更贵(应降价↓)", dn),
        ("  本店更便宜(应涨价↑)", up),
        ("  价格相同(—)", sq),
        ("[!]亏本预警(调价后低于采购价)", loss_cnt),
        ("需复核(比值异常)", rev_cnt),
        ("", ""),
        ("调价方案", "方案1: 本店售价 = 竞品售价 - 0.05元"),
        ("规格校验规则", "克重/容积差≤10可通过；数量不同排除；套装排除；单位不同排除"),
        ("方向说明", "↓=降价  ↑=涨价  —=不变"),
        ("亏本预警", "调价后价格 < 采购价 -> !!"),
        ("复核标记", "比值<0.5或>2.0 → 标注需复核"),
    ]
    for i, (k, v) in enumerate(stat, 1):
        ws2.cell(i, 1, k)
        ws2.cell(i, 2, v)
    for c in [1, 2]:
        cell = ws2.cell(1, c)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
    # 高亮关键行
    for ri in [7, 8, 9, 10, 11, 12]:
        for c in [1, 2]:
            ws2.cell(ri, c).fill = PatternFill("solid", fgColor="EBF3FB")
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 52

    # ── 品类明细 sheet ──────────────────────────────────────
    ws3 = wb.create_sheet("品类匹配明细")
    ws3_headers = ["品类", "本店品数", "竞品品数", "匹配数", "匹配率",
                   "降价↓", "涨价↑", "不变", "亏本预警", "需复核"]
    for ci, h in enumerate(ws3_headers, 1):
        cell = ws3.cell(1, ci, h)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        ws3.column_dimensions[get_column_letter(ci)].width = 14

    for cat in all_cats:
        lx_c = lx_by_cat.get(cat, [])
        cp_c = cp_by_cat.get(cat, [])
        matched = [r for r in results if r["lx_cat"] == cat]
        rate = len(matched)/len(lx_c)*100 if lx_c else 0
        dn_c = sum(1 for r in matched if r["direction"].startswith("↓"))
        up_c = sum(1 for r in matched if r["direction"].startswith("↑"))
        sq_c = sum(1 for r in matched if r["direction"].startswith("—"))
        loss_c = sum(1 for r in matched if r["loss"])
        rev_c = sum(1 for r in matched if r["review"])
        ws3.append([cat, len(lx_c), len(cp_c), len(matched),
                    f"{rate:.1f}%", dn_c, up_c, sq_c, loss_c, rev_c])

    last = ws3.max_row + 1
    ws3.cell(last, 1, "合计")
    ws3.cell(last, 2, sum(len(lx_by_cat[c]) for c in all_cats))
    ws3.cell(last, 3, sum(len(cp_by_cat[c]) for c in all_cats))
    ws3.cell(last, 4, len(results))
    ws3.cell(last, 5, f"{len(results)/len(lx_list)*100:.1f}%" if lx_list else "0%")
    ws3.cell(last, 6, dn)
    ws3.cell(last, 7, up)
    ws3.cell(last, 8, sq)
    ws3.cell(last, 9, loss_cnt)
    ws3.cell(last, 10, rev_cnt)
    for c in range(1, 11):
        cell = ws3.cell(last, c)
        cell.fill = PatternFill("solid", fgColor="D6E4F0")
        cell.font = Font(name="微软雅黑", bold=True)

    # ── 被排除商品 sheet ────────────────────────────────────
    ws4 = wb.create_sheet("规格校验排除记录")
    ws4_headers = ["本店品名", "本店规格", "竞品品名", "竞品规格",
                   "本店价格", "竞品价格", "排除原因", "品类"]
    for ci, h in enumerate(ws4_headers, 1):
        cell = ws4.cell(1, ci, h)
        cell.fill = PatternFill("solid", fgColor="808080")
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws4.column_dimensions["A"].width = 38
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 38
    ws4.column_dimensions["D"].width = 16
    ws4.column_dimensions["E"].width = 10
    ws4.column_dimensions["F"].width = 10
    ws4.column_dimensions["G"].width = 22
    ws4.column_dimensions["H"].width = 12
    for ri, r in enumerate(rejected, 2):
        ws4.cell(ri, 1, r["lx_name"])
        ws4.cell(ri, 2, r["lx_spec"])
        ws4.cell(ri, 3, r["cp_name"])
        ws4.cell(ri, 4, r["cp_spec"])
        ws4.cell(ri, 5, r["lx_price"])
        ws4.cell(ri, 6, r["cp_price"])
        ws4.cell(ri, 7, r["reason"])
        ws4.cell(ri, 8, r["cat"])
        for ci in range(1, 9):
            ws4.cell(ri, ci).border = border
            ws4.cell(ri, ci).font = Font(name="微软雅黑", size=10)
            ws4.cell(ri, ci).alignment = Alignment(horizontal="left" if ci in (1,3,7) else "center",
                                                     vertical="center")
        # 排除原因颜色
        ws4.cell(ri, 7).font = Font(name="微软雅黑", size=10, color="C00000")

    wb.save(out_path)
    print(f"\n>>> 输出: {out_path}")
    print(f">>> 通过: {len(results)} | 排除: {len(rejected)}")
    print(f">>> 降: {dn} | 涨: {up} | 不变: {sq}")
    print(f">>> 亏本: {loss_cnt} | 复核: {rev_cnt}")
    return len(results), len(rejected), dn, up, loss_cnt, rev_cnt


# ============================================================
# 执行两份比价
# ============================================================

# 1. 广安店 vs 小柴购
r1 = run_compare(
    lx_path   = r"C:\Users\13522\Desktop\鹿小仓广安店\鹿小仓广安店_库存合并总表.xlsx",
    comp_path = r"C:\Users\13522\Desktop\广安商超\小柴购_全量数据表_v5.xlsx",
    out_path  = r"C:\Users\13522\Desktop\鹿小仓广安店\调价方案1_鹿小仓广安店vs小柴购_比价表_v2.xlsx",
    lx_label  = "鹿小仓广安店",
    comp_label = "小柴购",
)

# 2. 财富店 vs 丽晨超市
r2 = run_compare(
    lx_path   = r"C:\Users\13522\Desktop\鹿小仓财富店\鹿小仓财富店_库存合并总表.xlsx",
    comp_path = r"C:\Users\13522\Desktop\厉臣超市\厉臣超市_标准化总数据表.xlsx",
    out_path  = r"C:\Users\13522\Desktop\鹿小仓财富店\调价方案1_鹿小仓财富店vs丽晨超市_比价表_v2.xlsx",
    lx_label  = "鹿小仓财富店",
    comp_label = "丽晨超市",
)

print("\n" + "="*60)
print("  汇总")
print("="*60)
print(f"广安店 vs 小柴购:  通过{r1[0]} 排除{r1[1]} 降{r1[2]} 涨{r1[3]} 亏本{r1[4]} 复核{r1[5]}")
print(f"财富店 vs 丽晨超市: 通过{r2[0]} 排除{r2[1]} 降{r2[2]} 涨{r2[3]} 亏本{r2[4]} 复核{r2[5]}")
