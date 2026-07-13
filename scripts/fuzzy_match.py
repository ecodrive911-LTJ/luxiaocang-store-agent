# -*- coding: utf-8 -*-
"""
鹿小仓 vs 小柴购 完整比对 + 输出
策略：品类过滤 → Jaccard关键词相似度 → difflib.SequenceMatcher
"""
import openpyxl, re, sys, math
from collections import defaultdict
import difflib

LUXIAO_PATH   = r"C:\Users\13522\Desktop\鹿小仓广安店\鹿小仓广安店_库存合并总表.xlsx"
XIAOCHAI_PATH = r"C:\Users\13522\Desktop\广安商超\小柴购_全量数据表_v5.xlsx"
OUTPUT_PATH   = r"C:\Users\13522\Desktop\鹿小仓广安店\调价方案1_鹿小仓vs小柴购_比价表.xlsx"

# ── 读取原始数据 ────────────────────────────────────────────
wb_lx = openpyxl.load_workbook(LUXIAO_PATH, data_only=True)
wb_xc = openpyxl.load_workbook(XIAOCHAI_PATH, data_only=True)
ws_lx, ws_xc = wb_lx.active, wb_xc.active

# 停用词（太常见无区分度）
STOPWORDS = {"的", "和", "与", "或", "及", "为", "比", "型", "装", "个",
             "包", "袋", "瓶", "支", "盒", "箱", "片", "卷", "米", "ml",
             "g", "kg", "cm", "mm", "plus", "版", "新款", "老款", "随机",
             "同款", "家用", "便携", "迷你", "小型", "大号", "中号", "小号"}

def clean_tokens(s):
    if not s:
        return set()
    s = str(s).strip()
    s = re.sub(r'[,，.。/／\-\(\)（）、\[\]【】]+', ' ', s)
    s = re.sub(r'\d+\s*(ml|g|kg|cm|mm|个|包|袋|瓶|支|盒|箱|片|卷|米)?', '', s)
    tokens = [t.strip().lower() for t in s.split() if len(t.strip()) >= 2 and t.strip() not in STOPWORDS]
    return set(tokens)

def jaccard(s1, s2):
    if not s1 or not s2:
        return 0.0
    inter = len(s1 & s2)
    union = len(s1 | s2)
    return inter / union if union > 0 else 0.0

def seq_ratio(a, b):
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a[:50], b[:50]).ratio()

# ── 加载鹿小仓 ─────────────────────────────────────────────
lx_list = []
for row in ws_lx.iter_rows(min_row=2, values_only=True):
    if not row or not row[1]:
        continue
    name = str(row[1]).strip()
    lx_list.append({
        "seq":   row[0],
        "name":  name,
        "cat":   (row[2] or "未分类").strip(),
        "spec":  row[3],
        "price": row[4],
        "cost":  row[5],
        "bar":   row[8],
        "tokens": clean_tokens(name),
    })

# ── 加载小柴购（过滤噪声）────────────────────────────────
NOISE = {"估配送费","红包立减","扫码领红包","红包满减","红包可用",
         "加购","立即购买","收藏","加入购物车","下单立减",
         "叠加优惠","满减券","领券","返利","优惠券"}
xc_list = []
for row in ws_xc.iter_rows(min_row=2, values_only=True):
    if not row or not row[1]:
        continue
    name = str(row[1]).strip()
    if len(name) < 4:
        continue
    if any(ns in name for ns in NOISE):
        continue
    price = row[5]
    if price and (float(price) <= 0 or float(price) > 2000):
        continue
    xc_list.append({
        "seq":   row[0],
        "name":  name,
        "cat":   (row[2] or "未分类").strip(),
        "spec":  row[4],
        "price": row[5],
        "tokens": clean_tokens(name),
    })

print(f"LX: {len(lx_list)}, XC: {len(xc_list)}")

# ── 按品类分组 ─────────────────────────────────────────────
lx_by_cat  = defaultdict(list)
xc_by_cat  = defaultdict(list)
for d in lx_list:
    lx_by_cat[d["cat"]].append(d)
for d in xc_list:
    xc_by_cat[d["cat"]].append(d)

# ── 匹配参数 ───────────────────────────────────────────────
JACCARD_THR   = 0.12   # Jaccard阈值（降低以提高召回）
SEQ_THR       = 0.45   # SequenceMatcher阈值（低Jaccard时备用）

# ── 执行匹配 ───────────────────────────────────────────────
results = []
match_counts = {"lx": 0, "xc_used": set()}

for cat, lx_items in lx_by_cat.items():
    xc_items = xc_by_cat.get(cat, [])
    if not xc_items:
        continue

    for lx in lx_items:
        best_score   = 0.0
        best_xc_name = None
        best_xc_price = None
        best_xc_spec  = None

        for xc in xc_items:
            # 跳过已匹配（取最优）
            # 跳过双方价格都为空的
            if lx["price"] is None and xc["price"] is None:
                continue

            j = jaccard(lx["tokens"], xc["tokens"])
            s = seq_ratio(lx["name"], xc["name"]) if j < JACCARD_THR else j

            # 通过任一阈值
            if j >= JACCARD_THR or s >= SEQ_THR:
                score = max(j, s)
                if score > best_score:
                    best_score   = score
                    best_xc_name  = xc["name"]
                    best_xc_price = xc["price"]
                    best_xc_spec  = xc["spec"]

        if best_score > 0 and best_xc_price is not None:
            lx_p_raw = lx["price"]
            xc_p_raw = best_xc_price
            lx_p = None
            xc_p = None
            try:
                if lx_p_raw is not None and str(lx_p_raw).strip() != '':
                    lx_p = float(lx_p_raw)
                if xc_p_raw is not None and str(xc_p_raw).strip() != '':
                    xc_p = float(xc_p_raw)
            except (ValueError, TypeError):
                pass

            ratio = (xc_p / lx_p) if (lx_p and xc_p and lx_p != 0) else None
            adj = round(xc_p - 0.05, 2) if xc_p is not None else None
            if adj is not None and adj < 0:
                adj = 0.0
            diff = round(adj - lx_p, 2) if (adj is not None and lx_p is not None) else None

            results.append({
                "lx_name":  lx["name"],
                "lx_cat":   lx["cat"],
                "lx_spec":  lx["spec"],
                "lx_price": lx_p,
                "lx_cost":  lx["cost"],
                "lx_bar":   lx["bar"],
                "xc_name":  best_xc_name,
                "xc_spec":  best_xc_spec,
                "xc_price": xc_p,
                "ratio":    ratio,
                "adj":      adj,
                "diff":     diff,
                "score":    best_score,
            })
            match_counts["lx"] += 1

print(f"\n匹配成功: {match_counts['lx']} / {len(lx_list)} LX商品")
# 按匹配分排序
results.sort(key=lambda x: x["score"], reverse=True)

# 显示头部样本
print("\n样本（前5条）:")
for r in results[:5]:
    print(f"  [{r['score']:.2f}] LX={r['lx_name'][:30]}")
    print(f"        XC={r['xc_name'][:30]}")
    ratio_str = f"{r['ratio']:.4f}" if r['ratio'] is not None else 'N/A'
    print(f"        ratio={ratio_str}  adj={r['adj']}")

# ── 统计 ────────────────────────────────────────────────────
comparable = [r for r in results if r["ratio"] is not None]
cheaper_than_xc = sum(1 for r in comparable if r["ratio"] and r["ratio"] < 1)
expensive_than_xc = sum(1 for r in comparable if r["ratio"] and r["ratio"] > 1)
equal = sum(1 for r in comparable if r["ratio"] and abs(r["ratio"] - 1) < 0.001)
print(f"\n统计（可比价商品 {len(comparable)} 个）:")
print(f"  鹿小仓比小柴购贵: {expensive_than_xc} 个")
print(f"  鹿小仓比小柴购便宜: {cheaper_than_xc} 个")
print(f"  价格相同: {equal} 个")
