# -*- coding: utf-8 -*-
"""
v3: 需复核记录深度重审
对v2中标注"需复核"的记录，用更严格的方法重新判定：
1. 品名核心词提取（品牌+品名）→ 必须品牌一致或明确可替代
2. 从品名中提取规格（不仅是规格列）→ 交叉验证
3. 数量/容量归一化 → 计算单价（元/ml 或 元/g）→ 比单价
4. 不同品类不可比（如白酒 vs 啤酒，面膜 vs 湿巾）
5. 最终分类：✅确认可比 / ❌误匹配 / ⚠需人工判断
"""
import openpyxl, re
from collections import Counter

def extract_brands(name):
    """提取品名中的品牌词"""
    if not name: return set()
    name = str(name)
    # 常见品牌
    brands = [
        "农夫山泉","可口可乐","百事","康师傅","统一","娃哈哈","雪花","百威","喜力",
        "奥利奥","好丽友","旺仔","喜之郎","德芙","杰士邦","杜蕾斯","冈本","第6感","名流",
        "清扬","滴露","南孚","得力","吉列","网易严选","李宁","惠普","闪迪","晶华",
        "薇婷","老干妈","珍极","醋天立","大大","趣多多","脉动","尖叫","芬达","红牛",
        "红星","牛栏山","古井贡","七道粮","草原白","九龙醉","徐九经","康保","龙岩",
        "衡水","北京二锅头","江小白","锐澳","RIO","宝娜斯","德佑","怡飘","小鹿妈妈",
        "老管家","柴选","高颅顶","胖MM","惠选","贝特幂","嘉士伯","立白","蓝漂",
        "佳邦士","兔之力","米菲娜","原生宠爱","黛曼普","稚优泉","姚记","瑞禾",
        "夏桐","燕京","翰思","金达日美","帆布","卡皮巴拉","太太乐","永和",
    ]
    found = set()
    for b in brands:
        if b in name:
            found.add(b)
    return found

def extract_spec_from_name(name):
    """从品名中提取规格信息"""
    if not name: return {}
    name = str(name)
    specs = {}

    # 克重
    m = re.search(r'(\d+(?:\.\d+)?)\s*(kg|g|克)', name, re.IGNORECASE)
    if m:
        val = float(m.group(1))
        specs['weight_g'] = val * 1000 if m.group(2).lower() in ('kg',) else val

    # 容积
    m = re.search(r'(\d+(?:\.\d+)?)\s*(ml|L|升|毫升)', name, re.IGNORECASE)
    if m:
        val = float(m.group(1))
        specs['volume_ml'] = val * 1000 if m.group(2).lower() in ('l', '升') else val

    # 数量
    m = re.search(r'(\d+)\s*(只|个|支|盒|袋|瓶|罐|听|包|条|双|对|副|张|片|份|把|块|贴|粒|副)', name)
    if m:
        specs['count'] = int(m.group(1))

    # 套装
    if re.search(r'[×x*]\s*\d+|整箱|组合装|礼盒|套装|半打|一打|到手\d+', name, re.IGNORECASE):
        specs['is_set'] = True

    # 乘法 ×N
    m = re.search(r'[×x*]\s*(\d+)', name, re.IGNORECASE)
    if m:
        specs['set_count'] = int(m.group(1))
        specs['is_set'] = True

    # "到手N瓶/罐/听"
    m = re.search(r'到手(\d+)\s*(瓶|罐|听|包|袋)', name)
    if m:
        specs['set_count'] = int(m.group(1))
        specs['is_set'] = True

    return specs

def get_unit_price(price, specs):
    """计算单价（元/g 或 元/ml）"""
    if not specs or not price:
        return None
    if 'weight_g' in specs:
        return price / specs['weight_g']
    if 'volume_ml' in specs:
        return price / specs['volume_ml']
    return None

def categorize_review(lx_name, lx_spec, lx_price, cp_name, cp_spec, cp_price):
    """
    对需复核记录重新判定
    返回: (verdict, reason)
    verdict: CONFIRM / REJECT / MANUAL
    """
    if not lx_name or not cp_name:
        return "REJECT", "品名缺失"

    lx_name_s = str(lx_name)
    cp_name_s = str(cp_name)

    # 1. 品牌校验
    lx_brands = extract_brands(lx_name_s)
    cp_brands = extract_brands(cp_name_s)

    # 如果双方都能提取到品牌，且品牌不同 → 判定不可比
    if lx_brands and cp_brands and not (lx_brands & cp_brands):
        # 特殊情况：竞品是"整箱"且本店也是"整箱"同一品牌的不同子品类
        return "REJECT", f"品牌不同({'/'.join(lx_brands)} vs {'/'.join(cp_brands)})"

    # 2. 从品名提取规格
    lx_ns = extract_spec_from_name(lx_name_s)
    cp_ns = extract_spec_from_name(cp_name_s)

    # 合并规格列和品名中的规格信息
    def merge_spec(col_spec, name_spec):
        result = dict(name_spec)
        if col_spec:
            if 'weight_g' not in result:
                m = re.search(r'(\d+(?:\.\d+)?)\s*(kg|g|克)', str(col_spec), re.IGNORECASE)
                if m:
                    val = float(m.group(1))
                    result['weight_g'] = val * 1000 if m.group(2).lower() in ('kg',) else val
            if 'volume_ml' not in result:
                m = re.search(r'(\d+(?:\.\d+)?)\s*(ml|L|升|毫升)', str(col_spec), re.IGNORECASE)
                if m:
                    val = float(m.group(1))
                    result['volume_ml'] = val * 1000 if m.group(2).lower() in ('l', '升') else val
        return result

    lx_full = merge_spec(lx_spec, lx_ns)
    cp_full = merge_spec(cp_spec, cp_ns)

    # 3. 套装 vs 单品
    lx_set = lx_full.get('is_set', False)
    cp_set = cp_full.get('is_set', False)

    # 如果一方是套装（如6瓶、12听、整箱）另一方是单品 → 比单价
    if lx_set != cp_set:
        # 尝试比单价
        lx_unit = get_unit_price(lx_price, lx_full)
        cp_unit = get_unit_price(cp_price, cp_full)
        if lx_unit and cp_unit:
            ratio = cp_unit / lx_unit if lx_unit else None
            if ratio and 0.3 <= ratio <= 3.0:
                return "MANUAL", f"套装vs单品-单价可比(竞品/本店单价={ratio:.2f})"
            elif ratio:
                return "REJECT", f"套装vs单品-单价差异过大(比值{ratio:.2f})"
        return "REJECT", "套装vs单品-无法计算单价"

    # 双方都是套装
    if lx_set and cp_set:
        lx_sc = lx_full.get('set_count', 0)
        cp_sc = cp_full.get('set_count', 0)
        if lx_sc and cp_sc and lx_sc != cp_sc:
            # 数量不同，比单价
            lx_unit = get_unit_price(lx_price, lx_full)
            cp_unit = get_unit_price(cp_price, cp_full)
            if lx_unit and cp_unit:
                ratio = cp_unit / lx_unit if lx_unit else None
                if ratio and 0.3 <= ratio <= 3.0:
                    return "MANUAL", f"套装数量不同({lx_sc}vs{cp_sc})-比单价(比值{ratio:.2f})"
                else:
                    return "REJECT", f"套装数量不同且单价差异过大"
            return "REJECT", f"套装数量不同({lx_sc}vs{cp_sc})"

    # 4. 品类差异检测（如白酒 vs 啤酒，零食 vs 面膜）
    category_pairs = [
        ({"白酒","二锅头","老白干","高粱","浓香","清香","酱香","兼香"}, {"啤酒","听装","听啤"}),
        ({"面膜","湿巾","湿纸巾"}, {"抽纸","纸巾","卷纸"}),
        ({"避孕套","安全套"}, {"避孕套","安全套"}),  # 同品类但要看品牌
        ({"洗发","护发"}, {"沐浴","沐浴露"}),
        ({"猫粮","猫砂","猫零食"}, {"犬粮","狗粮","犬零食"}),
        ({"可乐","汽水"}, {"果汁","果味"}),
    ]
    # 检测是否跨子品类
    for cat_a_words, cat_b_words in category_pairs:
        lx_in_a = any(w in lx_name_s for w in cat_a_words)
        lx_in_b = any(w in lx_name_s for w in cat_b_words)
        cp_in_a = any(w in cp_name_s for w in cat_a_words)
        cp_in_b = any(w in cp_name_s for w in cat_b_words)
        if (lx_in_a and cp_in_b) or (lx_in_b and cp_in_a):
            return "REJECT", f"子品类不同({'/'.join(cat_a_words)} vs {'/'.join(cat_b_words)})"

    # 5. 竞品价格异常检测（如0.1元、0.05元明显是异常价）
    if cp_price is not None and cp_price < 0.5:
        return "REJECT", f"竞品价格异常({cp_price}元)"
    if cp_price is not None and cp_price > 300 and lx_price is not None and lx_price < 50:
        return "REJECT", f"价格量级差异过大(本店{lx_price} vs 竞品{cp_price})"

    # 6. 规格列是品类名而非真实规格（如"矿泉水"、"饮料类"、"酒类"）
    # 这种情况下如果品名中能提取到规格，用品名规格比较
    spec_is_category = bool(re.match(r'^(矿泉水|饮料类?|酒类?|零食类?|食品饮料类?|日用百货|个护健康类?|'
                                     r'口腔护理|办公学习|五金家装|家庭清洁|生活用纸|品质百货|存储|外设|'
                                     r'狗粮|罐头|清凉一夏|居家待客|团建聚|复古怀旧|美发染发|汽车用品|'
                                     r'棋牌玩具|全部|花卉园艺|床上用品|拖鞋鞋类|出行无忧|品质百货.*|.*宠物.*)$',
                                     str(cp_spec).strip()))

    if spec_is_category:
        # 用品名提取的规格比较
        if ('weight_g' in lx_full and 'weight_g' in cp_full):
            diff = abs(lx_full['weight_g'] - cp_full['weight_g'])
            if diff > 10:
                return "REJECT", f"品名克重差{diff:.0f}g"
        elif ('volume_ml' in lx_full and 'volume_ml' in cp_full):
            diff = abs(lx_full['volume_ml'] - cp_full['volume_ml'])
            if diff > 10:
                return "REJECT", f"品名容积差{diff:.0f}ml"
        # 品名规格无法比较 → 检查品牌
        if lx_brands and cp_brands and (lx_brands & cp_brands):
            return "CONFIRM", "同品牌-竞品规格字段为品类名"
        # 如果品牌也提取不到，用品名相似度
        return "MANUAL", "竞品规格字段为品类名-需人工确认"

    # 7. 通过以上所有检查 → 确认可比
    return "CONFIRM", "通过深度复核"


# ============================================================
# 处理两份文件
# ============================================================

for path, label in [
    (r"C:\Users\13522\Desktop\鹿小仓广安店\调价方案1_鹿小仓广安店vs小柴购_比价表_v2.xlsx", "广安店vs小柴购"),
    (r"C:\Users\13522\Desktop\鹿小仓财富店\调价方案1_鹿小仓财富店vs丽晨超市_比价表_v2.xlsx", "财富店vs丽晨超市"),
]:
    print(f"\n{'='*60}")
    print(f"  {label} — 需复核深度重审")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["调价方案1_比价表"]

    verdicts = []
    for r in range(2, ws.max_row + 1):
        review = ws.cell(r, 16).value
        if not review or "复核" not in str(review):
            continue

        vals = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        verdict, reason = categorize_review(
            vals[2], vals[4], vals[5],  # name, spec, price (本店)
            vals[8], vals[9], vals[10],  # name, spec, price (竞品)
        )
        verdicts.append({
            "row": r,
            "lx_name": vals[2], "lx_spec": vals[4], "lx_price": vals[5],
            "cp_name": vals[8], "cp_spec": vals[9], "cp_price": vals[10],
            "ratio": vals[11], "adj": vals[12], "diff": vals[13],
            "direction": vals[14], "loss": vals[15],
            "verdict": verdict, "reason": reason,
        })

    # 统计
    v_counts = Counter(v["verdict"] for v in verdicts)
    print(f"\n总需复核: {len(verdicts)}")
    print(f"  CONFIRM (确认可比): {v_counts.get('CONFIRM', 0)}")
    print(f"  REJECT  (误匹配): {v_counts.get('REJECT', 0)}")
    print(f"  MANUAL  (需人工):  {v_counts.get('MANUAL', 0)}")

    # REJECT 原因分布
    reject_reasons = Counter(v["reason"] for v in verdicts if v["verdict"] == "REJECT")
    print(f"\nREJECT原因分布:")
    for reason, cnt in reject_reasons.most_common():
        print(f"  {reason}: {cnt}")

    MANUAL_reasons = Counter(v["reason"] for v in verdicts if v["verdict"] == "MANUAL")
    print(f"\nMANUAL原因分布:")
    for reason, cnt in MANUAL_reasons.most_common():
        print(f"  {reason}: {cnt}")

    # 打印CONFIRM的记录
    print(f"\n--- CONFIRM (确认可比) ---")
    for v in verdicts:
        if v["verdict"] == "CONFIRM":
            print(f"  {str(v['lx_name'])[:35]} | {v['lx_price']} vs {v['cp_price']} | {v['reason']}")

    # 打印MANUAL
    print(f"\n--- MANUAL (需人工判断) ---")
    for v in verdicts:
        if v["verdict"] == "MANUAL":
            print(f"  {str(v['lx_name'])[:35]} | {str(v['cp_name'])[:35]} | {v['reason']}")

    # 打印前10条REJECT
    print(f"\n--- REJECT (前15条) ---")
    for v in verdicts[:15]:
        if v["verdict"] == "REJECT":
            print(f"  {str(v['lx_name'])[:30]} vs {str(v['cp_name'])[:30]} | {v['reason']}")

    # 写入新的sheet
    wb2 = openpyxl.Workbook()
    ws_out = wb2.active
    ws_out.title = "需复核深度重审"

    headers = ["判定", "判定原因", "本店品名", "本店规格", "本店售价", "本店采购价",
               "竞品品名", "竞品规格", "竞品售价", "比值", "建议价", "差额", "方向", "亏本"]
    for ci, h in enumerate(headers, 1):
        cell = ws_out.cell(1, ci, h)
        cell.font = openpyxl.styles.Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 先CONFIRM，再MANUAL，最后REJECT
    order = {"CONFIRM": 0, "MANUAL": 1, "REJECT": 2}
    verdicts.sort(key=lambda x: (order.get(x["verdict"], 9), x["row"]))

    for ri, v in enumerate(verdicts, 2):
        vals = [v["verdict"], v["reason"], v["lx_name"], v["lx_spec"], v["lx_price"],
                None,  # 采购价 (需要从原表取)
                v["cp_name"], v["cp_spec"], v["cp_price"],
                v["ratio"], v["adj"], v["diff"], v["direction"], v["loss"]]
        for ci, val in enumerate(vals, 1):
            cell = ws_out.cell(ri, ci, val)
            cell.border = border
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="left" if ci in (3, 7) else "center",
                                        vertical="center")
        # 判定颜色
        color = {"CONFIRM": "375623", "REJECT": "C00000", "MANUAL": "FF6600"}[v["verdict"]]
        ws_out.cell(ri, 1).font = Font(name="微软雅黑", size=10, bold=True, color=color)
        ws_out.cell(ri, 2).font = Font(name="微软雅黑", size=10, color=color)

    widths = [8, 28, 38, 14, 9, 9, 38, 14, 9, 12, 10, 10, 8, 8]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws_out.column_dimensions[get_column_letter(i)].width = w

    # 统计sheet
    ws_stat = wb2.create_sheet("重审统计")
    ws_stat.cell(1, 1, "判定").font = Font(name="微软雅黑", bold=True)
    ws_stat.cell(1, 2, "数量").font = Font(name="微软雅黑", bold=True)
    ws_stat.cell(2, 1, "CONFIRM (确认可比)")
    ws_stat.cell(2, 2, v_counts.get('CONFIRM', 0))
    ws_stat.cell(3, 1, "REJECT (误匹配)")
    ws_stat.cell(3, 2, v_counts.get('REJECT', 0))
    ws_stat.cell(4, 1, "MANUAL (需人工)")
    ws_stat.cell(4, 2, v_counts.get('MANUAL', 0))
    ws_stat.cell(5, 1, "总计")
    ws_stat.cell(5, 2, len(verdicts))

    out_name = path.replace("_v2.xlsx", "_v2_复核重审.xlsx")
    wb2.save(out_name)
    print(f"\n>>> 输出: {out_name}")
